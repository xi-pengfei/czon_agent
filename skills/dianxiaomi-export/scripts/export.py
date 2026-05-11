#!/usr/bin/env python3
"""
店小蜜导单脚本

流程（单事务）：
  1. 锁定当前 pb_orders 中 order_status='Unlock' 的 id 集合
  2. SELECT * FROM v_dianxiaomi（视图自身已过滤 Unlock）
  3. 行级清洗：
       - 空邮编 → 000000
       - 备注正则 IOSS → 写入"卖家税号（IOSS）"
       - 非空邮编 → AI 批量校验城市/省/州
  4. 写 Excel → workspace/dianxiaomi_<ts>.xlsx
  5. UPDATE pb_orders SET order_status='Locked' WHERE id IN (...)
  6. COMMIT；任一步骤失败整事务回滚

依赖：pymysql、openpyxl、python-dotenv、pyyaml、openai
"""
from __future__ import annotations

import argparse
import json
import logging
import os
import re
import sys
import time
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import pymysql
import pymysql.cursors
import yaml
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import PatternFill

# ── 项目根目录定位（与 pledgebox-sync 一致）─────────────────
def find_root(start: Path) -> Path:
    for parent in [start, *start.parents]:
        if (parent / "main.py").exists():
            return parent
    return start

BASE_DIR = find_root(Path(__file__).resolve())
sys.path.insert(0, str(BASE_DIR))  # 让 core.llm 可被导入
load_dotenv(BASE_DIR / ".env")

from core.llm import LLM, PROVIDERS  # noqa: E402

# ── 日志 ────────────────────────────────────────────────
LOG_DIR = BASE_DIR / "logs"
LOG_DIR.mkdir(exist_ok=True)
log_file = LOG_DIR / f"dianxiaomi-export-{datetime.now():%Y%m%d}.log"
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[logging.FileHandler(log_file, encoding="utf-8"), logging.StreamHandler()],
)
log = logging.getLogger("dianxiaomi-export")

# ── 配置 ────────────────────────────────────────────────
DB_HOST = os.getenv("DB_HOST", "127.0.0.1")
DB_PORT = int(os.getenv("DB_PORT", 3306))
DB_USER = os.getenv("DB_USER", "root")
DB_PASS = os.getenv("DB_PASSWORD", "")
DB_NAME = os.getenv("DB_NAME", "nocobase")

WORKSPACE = BASE_DIR / "workspace"
WORKSPACE.mkdir(exist_ok=True)

# 关键列名（与视图列头对齐）
COL_ZIP   = "*邮编"
COL_CITY  = "*城市"
COL_STATE = "*省/州"
COL_CC    = "*国家二字码"
COL_NOTE  = "买家备注"
COL_IOSS  = "卖家税号（IOSS）"

AI_BATCH_SIZE = 20
EMPTY_ZIP_FILL = "000000"
HIGHLIGHT_FILL = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
DETAIL_PRINT_LIMIT = 50  # 控制台单类明细最多打印多少条；超过的写日志

# 用于在控制台引用每行的可读字段
COL_ORDER_NO = "*订单号"
COL_BUYER    = "*买家姓名"

# 税号识别：IOSS = IM + 10 数字（欧盟 IOSS 标准格式）
IOSS_RE = re.compile(r"\b(IM\d{10})\b", re.IGNORECASE)
# AI 兜底关键词（命中则把整段 note 丢给 AI 二次确认）
TAX_HINT_RE = re.compile(r"(IOSS|VAT|EORI|税号|TAX\s*ID|GST|ABN)", re.IGNORECASE)


# ════════════════════════════════════════════════════════════
#  修改追踪
# ════════════════════════════════════════════════════════════
class ChangeLog:
    """记录每一处单元格修改，用于：① 涂黄高亮 ② 控制台明细输出"""

    KINDS = ("zip_fill", "ioss_regex", "ioss_ai", "address_ai")

    def __init__(self):
        self.entries: List[Tuple[int, str, object, object, str]] = []

    def add(self, row_idx: int, col_name: str, old, new, kind: str) -> None:
        self.entries.append((row_idx, col_name, old, new, kind))

    def by_kind(self, kind: str):
        return [e for e in self.entries if e[4] == kind]

    def cells_to_highlight(self, headers: List[str]):
        """返回 [(excel_row_1based, excel_col_1based)]，表头占第 1 行"""
        col_idx = {h: i + 1 for i, h in enumerate(headers)}
        return [
            (r + 2, col_idx[c])
            for r, c, *_ in self.entries
            if c in col_idx
        ]


# ════════════════════════════════════════════════════════════
#  工具
# ════════════════════════════════════════════════════════════
def get_conn():
    return pymysql.connect(
        host=DB_HOST, port=DB_PORT,
        user=DB_USER, password=DB_PASS, database=DB_NAME,
        charset="utf8mb4",
        cursorclass=pymysql.cursors.DictCursor,
        autocommit=False,
    )


def load_llm() -> Optional[LLM]:
    """按 config.yaml 的 active_provider 构建 LLM；缺 key 时返回 None"""
    cfg_path = BASE_DIR / "config.yaml"
    if not cfg_path.exists():
        log.warning("config.yaml 不存在，跳过 AI")
        return None
    cfg = yaml.safe_load(cfg_path.read_text(encoding="utf-8")) or {}
    provider = cfg.get("active_provider", "kimi")
    pcfg = PROVIDERS.get(provider)
    if not pcfg:
        log.warning(f"未知 provider={provider}，跳过 AI")
        return None
    api_key = os.getenv(pcfg["env_key"], "")
    if not api_key:
        log.warning(f"未配置 {pcfg['env_key']}，跳过 AI")
        return None
    model = (cfg.get("providers", {}).get(provider) or {}).get("model")
    return LLM(provider=provider, api_key=api_key, model=model)


def parse_json_loose(text: str) -> Optional[object]:
    """从 LLM 返回中尽量解析出 JSON（去掉 ```json 包裹等）"""
    if not text:
        return None
    text = text.strip()
    # 去 markdown 代码块
    m = re.search(r"```(?:json)?\s*(.*?)```", text, re.DOTALL)
    if m:
        text = m.group(1).strip()
    try:
        return json.loads(text)
    except json.JSONDecodeError:
        # 尝试取首个 [...] 或 {...}
        for opener, closer in (("[", "]"), ("{", "}")):
            l, r = text.find(opener), text.rfind(closer)
            if l >= 0 and r > l:
                try:
                    return json.loads(text[l : r + 1])
                except json.JSONDecodeError:
                    continue
    return None


# ════════════════════════════════════════════════════════════
#  规则一：空邮编兜底
# ════════════════════════════════════════════════════════════
def fill_empty_zip(rows: List[Dict], changes: ChangeLog) -> int:
    n = 0
    for i, r in enumerate(rows):
        v = r.get(COL_ZIP)
        if v is None or str(v).strip() == "":
            changes.add(i, COL_ZIP, v, EMPTY_ZIP_FILL, "zip_fill")
            r[COL_ZIP] = EMPTY_ZIP_FILL
            n += 1
    return n


# ════════════════════════════════════════════════════════════
#  规则二：买家备注 → IOSS 税号
# ════════════════════════════════════════════════════════════
def extract_ioss_regex(rows: List[Dict], changes: ChangeLog) -> Tuple[int, List[int]]:
    """先用正则精确抓 IOSS。返回 (回填行数, 待 AI 复核的行号列表)"""
    filled = 0
    pending_ai: List[int] = []
    for i, r in enumerate(rows):
        if r.get(COL_IOSS):
            continue
        note = (r.get(COL_NOTE) or "").strip()
        if not note:
            continue
        m = IOSS_RE.search(note)
        if m:
            ioss = m.group(1).upper()
            changes.add(i, COL_IOSS, r.get(COL_IOSS), ioss, "ioss_regex")
            r[COL_IOSS] = ioss
            filled += 1
        elif TAX_HINT_RE.search(note):
            pending_ai.append(i)
    return filled, pending_ai


def ai_extract_ioss(llm: LLM, rows: List[Dict], indices: List[int], changes: ChangeLog) -> int:
    """把可疑行整段备注交给 AI，让它返回 IOSS。返回回填行数"""
    if not indices:
        return 0
    payload = [{"i": i, "note": rows[i].get(COL_NOTE) or ""} for i in indices]
    system = (
        "你是订单数据清洗助手。从买家备注中提取欧盟 IOSS 税号（格式：IM 开头 + 10 位数字）。"
        "只返回 JSON，无任何说明。格式 [{\"i\": 行号, \"ioss\": \"IM1234567890\" 或 \"\"}]。"
        "无法确认时返回空串。"
    )
    user = json.dumps(payload, ensure_ascii=False)
    filled = 0
    try:
        msg = llm.complete(system, [{"role": "user", "content": user}], [])
        data = parse_json_loose(msg.content or "")
        if isinstance(data, list):
            for item in data:
                if not isinstance(item, dict):
                    continue
                idx = item.get("i")
                ioss = (item.get("ioss") or "").strip().upper()
                if isinstance(idx, int) and 0 <= idx < len(rows) and IOSS_RE.match(ioss):
                    changes.add(idx, COL_IOSS, rows[idx].get(COL_IOSS), ioss, "ioss_ai")
                    rows[idx][COL_IOSS] = ioss
                    filled += 1
    except Exception as e:
        log.warning(f"AI 提取 IOSS 失败：{e}")
    return filled


# ════════════════════════════════════════════════════════════
#  规则三：AI 校验邮编 vs 城市/省州
# ════════════════════════════════════════════════════════════
def ai_fix_addresses(llm: LLM, rows: List[Dict], changes: ChangeLog) -> int:
    """对所有非空邮编行做批量校验，返回被修正的行数（按行计数，不按字段）"""
    targets: List[int] = []
    for i, r in enumerate(rows):
        zip_v = (r.get(COL_ZIP) or "").strip()
        if zip_v and zip_v != EMPTY_ZIP_FILL:
            targets.append(i)
    if not targets:
        return 0

    fixed_rows = 0
    for chunk_start in range(0, len(targets), AI_BATCH_SIZE):
        chunk = targets[chunk_start : chunk_start + AI_BATCH_SIZE]
        payload = [
            {
                "i": i,
                "country": rows[i].get(COL_CC) or "",
                "zip":     rows[i].get(COL_ZIP) or "",
                "city":    rows[i].get(COL_CITY) or "",
                "state":   rows[i].get(COL_STATE) or "",
            }
            for i in chunk
        ]
        system = (
            "你是国际地址校验助手。给定 [{i, country, zip, city, state}] 列表，"
            "判断每行邮编与城市/省州是否一致。"
            "若一致返回该行 {\"i\": i, \"ok\": true}；"
            "若不一致返回 {\"i\": i, \"ok\": false, \"city\": \"正确城市\", \"state\": \"正确省州\"}；"
            "若无法判断，返回 {\"i\": i, \"ok\": true}（保守不改）。"
            "只返回 JSON 数组，不要解释。城市/省州按当地常用拉丁拼写。"
        )
        user = json.dumps(payload, ensure_ascii=False)
        try:
            msg = llm.complete(system, [{"role": "user", "content": user}], [])
            data = parse_json_loose(msg.content or "")
        except Exception as e:
            log.warning(f"AI 地址校验失败（批次 {chunk_start}）：{e}")
            continue
        if not isinstance(data, list):
            continue
        for item in data:
            if not isinstance(item, dict) or item.get("ok", True):
                continue
            idx = item.get("i")
            if not (isinstance(idx, int) and 0 <= idx < len(rows)):
                continue
            new_city = (item.get("city") or "").strip()
            new_state = (item.get("state") or "").strip()
            row_changed = False
            old_city = rows[idx].get(COL_CITY) or ""
            old_state = rows[idx].get(COL_STATE) or ""
            if new_city and new_city != old_city:
                changes.add(idx, COL_CITY, old_city, new_city, "address_ai")
                rows[idx][COL_CITY] = new_city
                row_changed = True
            if new_state and new_state != old_state:
                changes.add(idx, COL_STATE, old_state, new_state, "address_ai")
                rows[idx][COL_STATE] = new_state
                row_changed = True
            if row_changed:
                fixed_rows += 1
    return fixed_rows


# ════════════════════════════════════════════════════════════
#  写 Excel（修改过的单元格涂黄）
# ════════════════════════════════════════════════════════════
def write_xlsx(rows: List[Dict], headers: List[str], changes: ChangeLog) -> Path:
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    out = WORKSPACE / f"dianxiaomi_{ts}.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "店小蜜导单"
    ws.append(headers)
    for r in rows:
        ws.append([r.get(h, "") for h in headers])

    # 给所有被修改过的单元格涂黄
    for excel_row, excel_col in changes.cells_to_highlight(headers):
        ws.cell(row=excel_row, column=excel_col).fill = HIGHLIGHT_FILL

    wb.save(out)
    return out


# ════════════════════════════════════════════════════════════
#  控制台明细打印
# ════════════════════════════════════════════════════════════
def _row_label(rows: List[Dict], i: int) -> str:
    """构造一行的可读标识：订单号 + 买家姓名 + 国家"""
    no   = rows[i].get(COL_ORDER_NO) or "?"
    name = rows[i].get(COL_BUYER) or ""
    cc   = rows[i].get(COL_CC) or ""
    suffix = f" ({name}, {cc})" if name else ""
    return f"{no}{suffix}"


def _print_section(title: str, lines: List[str]) -> None:
    if not lines:
        return
    print(f"\n● {title}（{len(lines)} 行）")
    shown = lines[:DETAIL_PRINT_LIMIT]
    for line in shown:
        print(f"    {line}")
    if len(lines) > DETAIL_PRINT_LIMIT:
        print(f"    … 余下 {len(lines) - DETAIL_PRINT_LIMIT} 行已写入日志")
        for line in lines[DETAIL_PRINT_LIMIT:]:
            log.info(f"[{title}] {line}")


def print_change_details(rows: List[Dict], changes: ChangeLog) -> None:
    # 1) 空邮编
    zip_lines = [_row_label(rows, e[0]) for e in changes.by_kind("zip_fill")]
    _print_section("空邮编 → 000000", zip_lines)

    # 2) IOSS 回填
    ioss_lines = []
    for e in changes.by_kind("ioss_regex"):
        ioss_lines.append(f"{_row_label(rows, e[0])}  →  {e[3]}  [正则]")
    for e in changes.by_kind("ioss_ai"):
        ioss_lines.append(f"{_row_label(rows, e[0])}  →  {e[3]}  [AI]")
    _print_section("回填 IOSS 税号", ioss_lines)

    # 3) AI 地址校正（按行聚合，避免一行 city+state 拆成两条）
    addr_by_row: Dict[int, List[Tuple[str, object, object]]] = {}
    for e in changes.by_kind("address_ai"):
        addr_by_row.setdefault(e[0], []).append((e[1], e[2], e[3]))
    addr_lines = []
    for i in sorted(addr_by_row.keys()):
        diffs = "; ".join(
            f"{col} '{old or ''}' → '{new}'" for col, old, new in addr_by_row[i]
        )
        addr_lines.append(f"{_row_label(rows, i)}  ({diffs})")
    _print_section("AI 校正地址", addr_lines)


def print_locked_orders(rows: List[Dict], target_ids: List[int]) -> None:
    print(f"\n● 锁定订单（{len(target_ids)} 单）")
    shown = rows[:DETAIL_PRINT_LIMIT]
    for i, r in enumerate(shown):
        print(f"    {_row_label(rows, i)}")
    if len(rows) > DETAIL_PRINT_LIMIT:
        print(f"    … 余下 {len(rows) - DETAIL_PRINT_LIMIT} 单已写入日志")
        for i in range(DETAIL_PRINT_LIMIT, len(rows)):
            log.info(f"[锁定订单] {_row_label(rows, i)}")


# ════════════════════════════════════════════════════════════
#  主流程
# ════════════════════════════════════════════════════════════
def main():
    parser = argparse.ArgumentParser(description="店小蜜导单（v_dianxiaomi → xlsx + 锁单）")
    parser.add_argument("--dry-run", action="store_true", help="只导出，不更新 order_status")
    parser.add_argument("--no-ai", action="store_true", help="跳过 AI 校验")
    parser.add_argument("--limit", type=int, default=0, help="仅处理前 N 条（0 = 全量）")
    args = parser.parse_args()

    t0 = time.time()
    print("=" * 52)
    print("  店小蜜导单")
    print("=" * 52)

    # 1) 连库
    try:
        conn = get_conn()
    except pymysql.Error as e:
        log.error(f"数据库连接失败：{e}")
        sys.exit(1)

    llm = None if args.no_ai else load_llm()
    out_path: Optional[Path] = None
    rollback_reason: Optional[str] = None
    stat = {"exported": 0, "ai_addr": 0, "zip_fill": 0, "ioss": 0, "locked": 0}

    try:
        with conn.cursor() as cur:
            # 2) 锁定 id 集合
            cur.execute("SELECT id FROM pb_orders WHERE order_status='Unlock'")
            id_rows = cur.fetchall()
            target_ids = [r["id"] for r in id_rows]
            if not target_ids:
                print("⚠ 当前 pb_orders 中没有 Unlock 订单，无需导出")
                conn.rollback()
                return

            # 3) 查视图
            cur.execute("SELECT * FROM v_dianxiaomi")
            rows = list(cur.fetchall())
            headers = [d[0] for d in cur.description]

        if args.limit and args.limit > 0:
            rows = rows[: args.limit]

        if not rows:
            print("⚠ v_dianxiaomi 返回 0 行，无需导出")
            conn.rollback()
            return

        stat["exported"] = len(rows)

        # 4) 行级清洗
        changes = ChangeLog()
        stat["zip_fill"] = fill_empty_zip(rows, changes)

        ioss_filled, pending_ai = extract_ioss_regex(rows, changes)
        stat["ioss"] += ioss_filled
        if llm and pending_ai:
            stat["ioss"] += ai_extract_ioss(llm, rows, pending_ai, changes)

        if llm:
            stat["ai_addr"] = ai_fix_addresses(llm, rows, changes)
        else:
            log.info("已跳过 AI 地址校验")

        # 5) 写 Excel（带高亮）
        out_path = write_xlsx(rows, headers, changes)
        log.info(f"写出 Excel：{out_path}")

        # 6) 锁单
        if args.dry_run:
            log.info("--dry-run：跳过 UPDATE")
            conn.rollback()
        else:
            with conn.cursor() as cur:
                placeholders = ",".join(["%s"] * len(target_ids))
                sql = f"UPDATE pb_orders SET order_status='Locked' WHERE id IN ({placeholders}) AND order_status='Unlock'"
                cur.execute(sql, target_ids)
                affected = cur.rowcount
                if affected != len(target_ids):
                    rollback_reason = f"UPDATE 影响 {affected} 行 ≠ 预期 {len(target_ids)} 行"
                    raise RuntimeError(rollback_reason)
                stat["locked"] = affected
            conn.commit()

    except Exception as e:
        conn.rollback()
        log.error(f"❌ 失败回滚：{e}")
        # 写盘已发生但事务回滚 → 删 Excel 防止误用
        if out_path and out_path.exists() and not args.dry_run:
            try:
                out_path.unlink()
                log.info(f"已删除未提交的 Excel：{out_path}")
            except OSError:
                pass
        sys.exit(2)
    finally:
        conn.close()

    # 7) 输出摘要
    elapsed = time.time() - t0
    print(f"\n{'=' * 52}")
    print(f"  [店小蜜导出完成]")
    print(f"{'=' * 52}")
    print(f"  导出订单数：{stat['exported']}")
    print(f"  AI 校正地址：{stat['ai_addr']} 行")
    print(f"  空邮编兜底：{stat['zip_fill']} 行")
    print(f"  回填 IOSS 税号：{stat['ioss']} 行")
    print(f"  锁定订单数：{stat['locked']}{'（dry-run 未锁）' if args.dry_run else ''}")
    print(f"  输出文件：{out_path.relative_to(BASE_DIR) if out_path else '-'}")
    print(f"  耗时：{elapsed:.1f} 秒")

    # 8) 修改明细 & 锁定订单清单
    if changes.entries:
        print(f"\n{'─' * 52}")
        print("  修改明细（Excel 中已用黄色背景标记）")
        print(f"{'─' * 52}")
        print_change_details(rows, changes)
    else:
        print("\n  无任何字段被修改")

    if not args.dry_run and stat["locked"] > 0:
        print(f"\n{'─' * 52}")
        print_locked_orders(rows, target_ids)
    print(f"{'=' * 52}")


if __name__ == "__main__":
    main()
